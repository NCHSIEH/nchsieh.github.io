#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VMH-EcoMap 主工具（Excel）→ 網站知識庫（JSON）轉檔器

用法：
    python3 tools/xlsx_to_json.py source/VMH-EcoMap-1_主工具_v1.x.xlsx

它會把 Excel 的每一個資料分頁轉成 data/*.json，網站直接讀這些 JSON。
Excel 仍然是唯一的「輸入處」，網站是唯一的「發布與歷程處」——
每跑一次轉檔、commit 一次，Git 就替你留下一版可回溯的歷程。

只依賴 openpyxl：  pip install openpyxl
"""

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("需要 openpyxl，請先執行：pip install openpyxl")


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"


# --------------------------------------------------------------------------
# 小工具
# --------------------------------------------------------------------------

def clean(value):
    """把儲存格值正規化成字串／數字／None。"""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        # Excel 常把整數存成 float，去掉沒有意義的 .0
        return int(value) if value.is_integer() else round(value, 4)
    if isinstance(value, str):
        text = value.strip()
        if text in ("", "None", "#N/A", "-"):
            return None
        return text
    return value


def split_list(value):
    """把 'A; B; C' 或 'A, B' 這種欄位切成陣列。"""
    if not value:
        return []
    text = str(value)
    parts = re.split(r"[;；,，、]\s*", text)
    return [p.strip() for p in parts if p.strip()]


def rows_of(ws, start_row, id_col=1, max_blank=40):
    """
    從 start_row 開始逐列讀取，跳過 ID 欄為空的列（Excel 預留的空白公式列）。
    連續 max_blank 列都沒資料就停，避免掃過整張表。
    """
    blank = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        key = clean(row[id_col - 1]) if len(row) >= id_col else None
        if key is None:
            blank += 1
            if blank >= max_blank:
                break
            continue
        blank = 0
        yield row


def cell(row, index):
    """依 0-based 索引安全取值。"""
    return clean(row[index]) if len(row) > index else None


def num(value, default=None):
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


# --------------------------------------------------------------------------
# 各分頁的轉檔器
# --------------------------------------------------------------------------

def parse_capabilities(wb):
    """03 能力總表 — 單一資料源，60+ 項能力。"""
    ws = wb["03_能力總表_Master"]
    out = []
    for row in rows_of(ws, 5):
        out.append({
            "id":            cell(row, 0),
            "domain":        cell(row, 1),
            "stage":         cell(row, 2),
            "bizLine":       cell(row, 3),
            "name":          cell(row, 4),
            "nameEn":        cell(row, 5),
            "outcome":       cell(row, 6),
            "puzzle":        cell(row, 7),
            "lead":          cell(row, 8),
            "candidate":     cell(row, 9),
            "gapNext":       cell(row, 10),
            "gates":         split_list(cell(row, 11)),
            "gateStatus":    cell(row, 12),
            "valueGov":      num(cell(row, 13)),
            "valueClinical": num(cell(row, 14)),
            "valueBusiness": num(cell(row, 15)),
            "value":         num(cell(row, 16)),
            "readyReg":      num(cell(row, 17)),
            "readyPartner":  num(cell(row, 18)),
            "readyFunding":  num(cell(row, 19)),
            "ready":         num(cell(row, 20)),
            "risk":          cell(row, 21),
            "decision":      cell(row, 22),
            "depends":       split_list(cell(row, 23)),
            "deliverable":   cell(row, 24),
            "evidence":      cell(row, 25),
            "updated":       cell(row, 26),
            "maturity":      cell(row, 27),
            "acceptanceKpi": cell(row, 28),
            "dataDeps":      cell(row, 29),
            "evidenceLevel": cell(row, 30),
            "opportunities": split_list(cell(row, 31)),
        })
    return out


def parse_partners(wb):
    """04 夥伴拼圖 — 名冊與七構面資格評分。"""
    ws = wb["04_夥伴拼圖_Partners"]
    out = []
    for row in rows_of(ws, 5, max_blank=6):
        pid = cell(row, 0)
        # 只取名冊列；下方的「夥伴 × 領域覆蓋矩陣」A 欄是「代號＋名稱」，格式不同
        if not pid or not re.fullmatch(r"P-\d+", str(pid)):
            continue
        domains = [f"D{n:02d}" for n in range(1, 11) if cell(row, 19 + n)]
        out.append({
            "id":        pid,
            "name":      cell(row, 1),
            "nameEn":    cell(row, 2),
            "type":      cell(row, 3),
            "country":   cell(row, 4),
            "domains":   domains or split_list(cell(row, 5)),
            "role":      cell(row, 6),
            "vnPresence": cell(row, 7),
            "scores": {
                "references":   num(cell(row, 8), 0),
                "advisory":     num(cell(row, 9), 0),
                "build":        num(cell(row, 10), 0),
                "operate":      num(cell(row, 11), 0),
                "finance":      num(cell(row, 12), 0),
                "interop":      num(cell(row, 13), 0),
                "localization": num(cell(row, 14), 0),
            },
            "fit":        num(cell(row, 15), 0),
            "sourceRisk": cell(row, 16),
            "status":     cell(row, 17),
            "nextAction": cell(row, 18),
            "label":      cell(row, 19) or f"{pid} {cell(row, 1)}",
        })
    return out


def parse_gates(wb):
    """05 法規門檻 — B 軸：越南強制法規＋國際標準。"""
    ws = wb["05_法規門檻_Gates"]
    stages = ["G0", "G1", "G2", "G3", "G4", "G5", "G6"]
    out = []
    for row in rows_of(ws, 5):
        applies = [s for i, s in enumerate(stages) if cell(row, 6 + i)]
        out.append({
            "id":        cell(row, 0),
            "category":  cell(row, 1),
            "title":     cell(row, 2),
            "authority": cell(row, 3),
            "level":     cell(row, 4),
            "status":    cell(row, 5),
            "stages":    applies,
            "owner":     cell(row, 13),
            "url":       cell(row, 14),
            "note":      cell(row, 15),
        })
    return out


def parse_workpackages(wb):
    """06 行動路線 — 20 個工作包與 60 個月時程。"""
    ws = wb["06_行動路線_Roadmap"]
    out = []
    for row in rows_of(ws, 5):
        out.append({
            "id":         cell(row, 0),
            "stage":      cell(row, 1),
            "name":       cell(row, 2),
            "lead":       cell(row, 3),
            "partners":   split_list(cell(row, 4)),
            "predecessors": split_list(cell(row, 5)),
            "startMonth": num(cell(row, 6)),
            "months":     num(cell(row, 7)),
            "endMonth":   num(cell(row, 8)),
            "status":     cell(row, 9),
            "percent":    num(cell(row, 10), 0),
            "acceptance": cell(row, 11),
            "caps":       cell(row, 12),
        })
    return out


def parse_risks(wb):
    """08 風險清冊 — 機率 × 衝擊。"""
    ws = wb["08_風險清冊_Risk"]
    out = []
    for row in rows_of(ws, 5):
        out.append({
            "id":         cell(row, 0),
            "category":   cell(row, 1),
            "statement":  cell(row, 2),
            "likelihood": num(cell(row, 3)),
            "impact":     num(cell(row, 4)),
            "score":      num(cell(row, 5)),
            "level":      cell(row, 6),
            "mitigation": cell(row, 7),
            "owner":      cell(row, 8),
            "status":     cell(row, 9),
            "trigger":    cell(row, 10),
            "links":      split_list(cell(row, 11)),
        })
    return out


def parse_kpis(wb):
    """09 KPI 計分卡 — 效益驗收指標。"""
    ws = wb["09_KPI計分卡_KPI"]
    out = []
    for row in rows_of(ws, 5):
        kid = cell(row, 0)
        # 表尾有一列「KPI 使用規則｜How to use」的說明，不是指標
        if not kid or not re.match(r"KPI-", str(kid)):
            continue
        out.append({
            "id":          kid,
            "perspective": cell(row, 1),
            "name":        cell(row, 2),
            "unit":        cell(row, 3),
            "direction":   cell(row, 4),
            "baseline":    num(cell(row, 5)),
            "target":      num(cell(row, 6)),
            "delta":       num(cell(row, 7)),
            "frequency":   cell(row, 8),
            "owner":       cell(row, 9),
            "dataSource":  cell(row, 10),
            "note":        cell(row, 11),
            "formula":     cell(row, 12),
            "numerator":   cell(row, 13),
            "denominator": cell(row, 14),
            "sourceSystem": cell(row, 15),
            "baselineNote": cell(row, 16),
        })
    return out


def parse_opportunities(wb):
    """15 商機字典 — 85 門可以獨立談的生意。"""
    ws = wb["15_商機字典_OppLibrary"]
    out = []
    for row in rows_of(ws, 5):
        oid = cell(row, 0)
        if not oid or not str(oid).startswith("OPP-"):
            continue
        out.append({
            "id":         oid,
            "name":       cell(row, 1),
            "nameEn":     cell(row, 2),
            "family":     cell(row, 3),
            "whatWeSell": cell(row, 4),
            "vnMaturity": cell(row, 5),
            "domains":    split_list(cell(row, 6)),
            "capCount":   num(cell(row, 7), 0),
            "avgValue":   num(cell(row, 8)),
            "avgReady":   num(cell(row, 9)),
            "gapCount":   num(cell(row, 10), 0),
            "position":   cell(row, 11),
            "stage":      cell(row, 12),
            "payer":      cell(row, 13),
            "txnMode":    cell(row, 14),
            "amountTier": cell(row, 15),
            "standalone": cell(row, 16),
            "lead":       cell(row, 17),
            "vnNotes":    cell(row, 18),
            "inPipeline": bool(num(cell(row, 19), 0)),
            "label":      cell(row, 20),
            "competitors": cell(row, 21),
            "ourAngle":   cell(row, 22),
            "leadCaps":   num(cell(row, 23), 0),
            "priority":   num(cell(row, 24)),
            "rank":       num(cell(row, 25)),
        })
    return out


def parse_pipeline(wb):
    """14 商機清單 — 現在正在談的生意。"""
    ws = wb["14_商機清單_Pipeline"]
    out = []
    for row in rows_of(ws, 5):
        label = cell(row, 0)
        match = re.match(r"(OPP-\d+)", str(label)) if label else None
        if not match:
            continue
        out.append({
            "oppId":      match.group(1),
            "name":       cell(row, 1),
            "family":     cell(row, 2),
            "vnMaturity": cell(row, 3),
            "domains":    split_list(cell(row, 4)),
            "capCount":   num(cell(row, 5), 0),
            "avgValue":   num(cell(row, 6)),
            "avgReady":   num(cell(row, 7)),
            "gapCount":   num(cell(row, 8), 0),
            "payer":      cell(row, 12) or cell(row, 9),
            "txnMode":    cell(row, 13) or cell(row, 10),
            "amountTier": cell(row, 14) or cell(row, 11),
            "dealStage":  cell(row, 15),
            "lead":       cell(row, 16),
            "nextAction": cell(row, 17),
            "note":       cell(row, 18),
        })
    return out


def parse_cap_opp_matrix(wb):
    """16 能力商機矩陣 — 能力 ↔ 商機的多對多對照。"""
    ws = wb["16_能力商機矩陣_CapOppMap"]
    header = [clean(c) for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    opp_cols = {i: h for i, h in enumerate(header) if h and str(h).startswith("OPP-")}
    mapping = {}
    for row in rows_of(ws, 5):
        cid = cell(row, 0)
        if not cid or not str(cid).startswith("CAP-"):
            continue
        mapping[cid] = [oid for i, oid in opp_cols.items() if cell(row, i)]
    return mapping


def parse_log(wb):
    """10 決策紀錄 — 來源清冊、決策紀錄、版本紀錄三段。"""
    ws = wb["10_決策紀錄_Log"]
    sources, decisions, versions = [], [], []
    section = None
    for row in ws.iter_rows(values_only=True):
        first = clean(row[0]) if row else None
        text = str(first) if first else ""
        if text.startswith("A."):
            section = "sources"; continue
        if text.startswith("B."):
            section = "decisions"; continue
        if text.startswith("C."):
            section = "versions"; continue
        if not first or text in ("ID", "版本"):
            continue
        if section == "sources" and text.startswith("SRC-"):
            sources.append({
                "id": first, "title": cell(row, 1), "org": cell(row, 2),
                "purpose": cell(row, 3), "url": cell(row, 4),
                "checked": cell(row, 5), "status": cell(row, 6), "note": cell(row, 7),
            })
        elif section == "decisions" and text.startswith("D-"):
            decisions.append({
                "id": first, "date": cell(row, 1), "topic": cell(row, 2),
                "decision": cell(row, 3), "by": cell(row, 4), "status": cell(row, 5),
            })
        elif section == "versions":
            versions.append({
                "version": str(first), "date": cell(row, 1), "audience": cell(row, 2),
                "scope": cell(row, 3), "note": cell(row, 4),
            })
    return sources, decisions, versions


def parse_scenario(wb):
    """02 情境權重 — 專案輪廓與評分權重。"""
    ws = wb["02_情境權重_Scenario"]
    profile, weights = [], []
    section = None
    for row in ws.iter_rows(values_only=True):
        first = clean(row[0]) if row else None
        text = str(first) if first else ""
        if text.startswith("A."):
            section = "profile"; continue
        if text.startswith("B."):
            section = "weights"; continue
        if text.startswith("C."):
            section = "baseline"; continue
        if not first or text.startswith("欄位") or text.startswith("構面"):
            continue
        entry = {"field": first, "value": cell(row, 1), "note": cell(row, 3)}
        if section == "profile":
            profile.append(entry)
        elif section == "weights":
            weights.append({"field": first, "value": cell(row, 1), "note": cell(row, 2) or cell(row, 3)})
    return profile, weights


def parse_portfolio(wb):
    """07 整案量級 — 成本驅動與交易模式比較。"""
    ws = wb["07_整案量級_Portfolio"]
    drivers, summary, models = [], [], []
    section = None
    for row in ws.iter_rows(values_only=True):
        first = clean(row[0]) if row else None
        text = str(first) if first else ""
        if text.startswith("A."):
            section = "drivers"; continue
        if text.startswith("B."):
            section = "summary"; continue
        if text.startswith("C."):
            section = "models"; continue
        if not first or text in ("驅動", "項目", "模式"):
            continue
        if section == "drivers":
            drivers.append({"name": first, "value": cell(row, 1), "unit": cell(row, 2), "note": cell(row, 3)})
        elif section == "summary":
            summary.append({"name": first, "value": cell(row, 1), "note": cell(row, 3)})
        elif section == "models":
            models.append({
                "name": first,
                "publicControl": num(cell(row, 1)), "speed": num(cell(row, 2)),
                "overall": num(cell(row, 3)), "privateCapital": num(cell(row, 4)),
                "revenueRisk": num(cell(row, 5)), "complexity": num(cell(row, 6)),
                "nearTermFit": num(cell(row, 7)), "note": cell(row, 8),
            })
    return {"drivers": drivers, "summary": summary, "models": models}


# --------------------------------------------------------------------------
# 主流程
# --------------------------------------------------------------------------

def write_json(name, payload):
    path = DATA_DIR / f"{name}.json"
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
    )
    size = path.stat().st_size
    count = len(payload) if isinstance(payload, (list, dict)) else 1
    print(f"  data/{name}.json  ({count} 筆 / {size // 1024} KB)")


def main():
    if len(sys.argv) < 2:
        candidates = sorted((ROOT / "source").glob("*.xlsx"))
        if not candidates:
            sys.exit("用法：python3 tools/xlsx_to_json.py <主工具.xlsx>")
        src = candidates[-1]
    else:
        src = Path(sys.argv[1])

    if not src.exists():
        sys.exit(f"找不到檔案：{src}")

    print(f"讀取：{src.name}")
    wb = openpyxl.load_workbook(src, data_only=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    capabilities = parse_capabilities(wb)
    partners = parse_partners(wb)
    gates = parse_gates(wb)
    workpackages = parse_workpackages(wb)
    risks = parse_risks(wb)
    kpis = parse_kpis(wb)
    opportunities = parse_opportunities(wb)
    pipeline = parse_pipeline(wb)
    cap_opp = parse_cap_opp_matrix(wb)
    sources, decisions, versions = parse_log(wb)
    profile, weights = parse_scenario(wb)
    portfolio = parse_portfolio(wb)

    write_json("capabilities", capabilities)
    write_json("partners", partners)
    write_json("gates", gates)
    write_json("workpackages", workpackages)
    write_json("risks", risks)
    write_json("kpis", kpis)
    write_json("opportunities", opportunities)
    write_json("pipeline", pipeline)
    write_json("cap-opp", cap_opp)
    write_json("sources", sources)
    write_json("decisions", decisions)
    write_json("versions", versions)
    write_json("portfolio", portfolio)

    latest = versions[-1] if versions else {}
    meta = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "sourceFile": src.name,
        "toolVersion": latest.get("version"),
        "toolVersionDate": latest.get("date"),
        "profile": profile,
        "weights": weights,
        "counts": {
            "capabilities": len(capabilities),
            "partners": len(partners),
            "gates": len(gates),
            "workpackages": len(workpackages),
            "risks": len(risks),
            "kpis": len(kpis),
            "opportunities": len(opportunities),
            "pipeline": len(pipeline),
            "decisions": len(decisions),
            "sources": len(sources),
        },
    }
    write_json("meta", meta)
    print("完成。接著 git add / commit，歷程就會留在版本庫裡。")


if __name__ == "__main__":
    main()
